#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PDF题库转Excel工具
用于将PDF格式的题库转换为Excel格式
"""

import sys
from pathlib import Path
from PDF题库解析 import PDFTikuParser

try:
    import openpyxl
except ImportError:
    print("错误：需要安装 openpyxl 库")
    print("运行: pip install openpyxl")
    sys.exit(1)

def pdf_to_excel(pdf_file, excel_file=None):
    """将PDF题库转换为Excel格式"""
    pdf_path = Path(pdf_file)
    
    if not pdf_path.exists():
        print(f"错误：文件不存在 - {pdf_file}")
        return False
    
    # 解析PDF文件
    print(f"正在解析PDF文件: {pdf_path.name}")
    print("提示：PDF解析可能需要较长时间，请耐心等待...")
    print("-" * 60)
    
    parser = PDFTikuParser(pdf_path)
    questions = parser.parse()
    
    if not questions:
        print("\n错误：未能从PDF文档中解析到题目")
        print("\n可能的原因：")
        print("1. PDF是扫描版（图片格式），无法提取文本")
        print("2. PDF格式不规范，无法识别题目结构")
        print("3. 需要安装pdfplumber库: pip install pdfplumber")
        print("\n建议：")
        print("- 如果是扫描版PDF，需要先进行OCR识别")
        print("- 尝试手动复制粘贴到Word或Excel")
        return False
    
    print(f"\n成功解析 {len(questions)} 道题目")
    
    # 创建Excel文件
    if excel_file is None:
        excel_file = pdf_path.with_suffix('.xlsx')
    else:
        excel_file = Path(excel_file)
    
    print(f"正在创建Excel文件: {excel_file.name}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题库"
    
    # 写入表头
    headers = ['序号', '题目', 'A', 'B', 'C', 'D', '答案', '题型', '解析']
    ws.append(headers)
    
    # 设置表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入题目
    for i, q in enumerate(questions, 1):
        row = [
            i,
            q.get('question', ''),
            q.get('options', {}).get('A', ''),
            q.get('options', {}).get('B', ''),
            q.get('options', {}).get('C', ''),
            q.get('options', {}).get('D', ''),
            q.get('answer', ''),
            q.get('type', ''),
            q.get('explanation', '')
        ]
        ws.append(row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 50
    
    # 设置文本自动换行
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 保存
    wb.save(excel_file)
    print(f"\n[成功] 转换完成！Excel文件已保存到: {excel_file}")
    print(f"共 {len(questions)} 道题目")
    
    # 显示题型统计
    type_stats = {}
    for q in questions:
        qtype = q.get('type', '未知')
        type_stats[qtype] = type_stats.get(qtype, 0) + 1
    
    print("\n题型分布：")
    for qtype, count in sorted(type_stats.items(), key=lambda x: x[1], reverse=True):
        print(f"  {qtype}: {count}题")
    
    print("\n提示：建议打开Excel文件检查转换结果，必要时手动调整")
    
    return True

def batch_convert():
    """批量转换当前目录下的所有PDF文件"""
    current_dir = Path(__file__).parent
    pdf_files = list(current_dir.glob('*.pdf'))
    
    if not pdf_files:
        print("当前目录没有PDF文件")
        return
    
    print(f"找到 {len(pdf_files)} 个PDF文件")
    print("=" * 60)
    
    success_count = 0
    for pdf_file in pdf_files:
        print(f"\n处理: {pdf_file.name}")
        print("-" * 60)
        if pdf_to_excel(pdf_file):
            success_count += 1
        print("=" * 60)
    
    print(f"\n批量转换完成！")
    print(f"成功: {success_count}/{len(pdf_files)}")

def main():
    """主函数"""
    print("=" * 60)
    print("  PDF题库转Excel工具")
    print("=" * 60)
    print()
    print("注意：")
    print("- PDF解析需要较长时间，请耐心等待")
    print("- 仅支持文本型PDF，扫描版PDF无法识别")
    print("- 建议转换后检查Excel内容，必要时手动调整")
    print()
    
    if len(sys.argv) > 1:
        # 命令行模式
        pdf_file = sys.argv[1]
        excel_file = sys.argv[2] if len(sys.argv) > 2 else None
        pdf_to_excel(pdf_file, excel_file)
    else:
        # 交互模式
        print("1. 转换单个PDF文件")
        print("2. 批量转换当前目录所有PDF文件")
        print("0. 退出")
        print()
        
        choice = input("请选择: ").strip()
        
        if choice == '1':
            pdf_file = input("\n请输入PDF文件名（含.pdf后缀）: ").strip()
            if not pdf_file:
                print("文件名不能为空")
            else:
                excel_file = input("请输入Excel文件名（直接回车则自动命名）: ").strip()
                excel_file = excel_file if excel_file else None
                pdf_to_excel(pdf_file, excel_file)
        elif choice == '2':
            confirm = input("\n批量转换可能需要较长时间，确认继续？(y/n): ").strip().lower()
            if confirm == 'y':
                batch_convert()
        
        print()
        input("按回车键退出...")

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n已取消")
    except Exception as e:
        print(f"\n发生错误: {e}")
        import traceback
        traceback.print_exc()
        input("按回车键退出...")

